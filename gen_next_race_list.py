# coding: UTF-8
import os
import time
import datetime

from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import openpyxl


POGPATH = os.getenv("HOMEDRIVE", "None") + os.getenv("HOMEPATH", "None") + "/Dropbox/POG/"
WBPATH = (POGPATH + "POG_HorseList.xlsx").replace("\\", "/")
HTMLPATH = (POGPATH + "next_race_list.html").replace("\\", "/")
WEBDRIVERPATH = r"C:\Selenium\chromedriver.exe"

NEXT_RACE_URL = "http://db.netkeiba.com/community/?pid=horse_info_next&id={}"

mytoday = datetime.date.today()

options = Options()
options.add_argument('--headless')
driver = webdriver.Chrome(executable_path=WEBDRIVERPATH, options=options)


def main():
    global driver

    out_next_race_html(get_next_race_list())

    driver.quit()


def out_next_race_html(next_race_list):
    f = open(HTMLPATH, mode="w", encoding="utf-8")
    f.write('<table>\n')
    for row in next_race_list:
        horse_owner_name = row[1] + row[0]
        race = row[8] + "(" + row[7] + ") new!" if row[9] == "new" else row[8] + "(" + row[7] + ")"
        f.write("<tr><td>" + horse_owner_name + "</td><td>" + race + "</td></tr>\n")
    f.write("</table>\n")

    f.close()


def get_next_race_list():
    wb = openpyxl.load_workbook(WBPATH)
    wshl = wb["POHorseList"]

    xlrow = 1
    while wshl.cell(row=xlrow, column=6).value == "-":

        horse_id = wshl.cell(row=xlrow, column=7).value
        next_race_date = wshl.cell(row=xlrow, column=8).value
        next_race_name = wshl.cell(row=xlrow, column=9).value

        next_race = get_next_race(horse_id)
        if not next_race[0]:
            wshl.cell(row=xlrow, column=8).value = "-"
            wshl.cell(row=xlrow, column=9).value = "-"
            wshl.cell(row=xlrow, column=10).value = "-"
            xlrow += 1
            continue
        nrd_new = datetime.datetime.strptime(next_race[0], "%Y/%m/%d")

        if (nrd_new.date() - mytoday).days <= 0:
            wshl.cell(row=xlrow, column=8).value = "-"
            wshl.cell(row=xlrow, column=9).value = "-"
            wshl.cell(row=xlrow, column=10).value = "-"
            xlrow += 1
            continue

        if next_race_date == next_race[0] and next_race_name == next_race[1]:
            wshl.cell(row=xlrow, column=10).value = "-"
        else:
            wshl.cell(row=xlrow, column=8).value = next_race[0]
            wshl.cell(row=xlrow, column=9).value = next_race[1]
            wshl.cell(row=xlrow, column=10).value = "new"
        xlrow += 1

    wb.save(WBPATH)

    next_race_list_all = [[cell.value for cell in row] for row in wshl["A1:J" + str(xlrow - 1)]]
    return [row for row in next_race_list_all if row[7] != "-"]


def get_next_race(horse_id):
    global driver
    global NEXT_RACE_URL

    print(horse_id + "start")
    time.sleep(1)
    driver.get(NEXT_RACE_URL.format(horse_id))
    print(horse_id + "driver.get")
    html = driver.page_source.encode('utf-8')
    print(horse_id + "driver.page_source.encode")
    soup = BeautifulSoup(html, 'html.parser')

    next_race_dt = soup.find("dt", string="次走予定")
    if not next_race_dt:
        return [None, None]
    next_race_dd = next_race_dt.find_next("dd")
    if not next_race_dd:
        return [None, None]

    return [next_race_dd.contents[0].strip(), next_race_dd.contents[1].contents[0]]


if __name__ == "__main__":
    main()
